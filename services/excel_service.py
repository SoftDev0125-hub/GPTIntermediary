"""
Excel Service - Handle Excel spreadsheet operations
"""
import os
import logging
from typing import Dict, Any, List, Optional
from datetime import datetime

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Try to import openpyxl
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Fill, PatternFill, Border, Side, Alignment
    from openpyxl.utils import get_column_letter, column_index_from_string
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    logger.warning("openpyxl not installed. Install with: pip install openpyxl")


class ExcelService:
    """Service for handling Excel spreadsheet operations"""
    
    def __init__(self):
        """Initialize the Excel service"""
        if not HAS_OPENPYXL:
            logger.warning("Excel service initialized without openpyxl library")
    
    async def create_spreadsheet(self, file_path: str, sheet_name: str = "Sheet1") -> Dict[str, Any]:
        """
        Create a new Excel spreadsheet
        
        Args:
            file_path: Path where the spreadsheet should be saved
            sheet_name: Name of the first sheet
        
        Returns:
            Dictionary with operation result
        """
        if not HAS_OPENPYXL:
            return {
                "success": False,
                "error": "openpyxl library not installed"
            }
        
        try:
            # Ensure directory exists
            directory = os.path.dirname(file_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Ensure file has .xlsx extension
            if not file_path.lower().endswith('.xlsx'):
                file_path = file_path + '.xlsx'
            
            # Create new workbook
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            # Save the workbook
            wb.save(file_path)
            
            logger.info(f"Created Excel spreadsheet: {file_path}")
            return {
                "success": True,
                "file_path": file_path,
                "sheet_name": sheet_name,
                "message": f"Spreadsheet created successfully"
            }
        except Exception as e:
            logger.error(f"Error creating spreadsheet: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def open_spreadsheet(self, file_path: str) -> Dict[str, Any]:
        """
        Open an existing Excel spreadsheet
        
        Args:
            file_path: Path to the spreadsheet
        
        Returns:
            Dictionary with spreadsheet data
        """
        if not HAS_OPENPYXL:
            return {
                "success": False,
                "error": "openpyxl library not installed"
            }
        
        try:
            if not os.path.exists(file_path):
                return {
                    "success": False,
                    "error": f"Spreadsheet not found: {file_path}"
                }
            
            # Load workbook
            wb = load_workbook(file_path)
            
            # Get all sheet names
            sheet_names = wb.sheetnames
            
            # Get data from active sheet
            ws = wb.active
            data = []
            
            # Read all rows (limit to first 1000 rows for performance)
            for row_idx, row in enumerate(ws.iter_rows(max_row=1000, values_only=False), start=1):
                row_data = []
                for cell in row:
                    # Helper function to convert RGB to string
                    def rgb_to_string(rgb_obj):
                        if rgb_obj is None:
                            return None
                        try:
                            # RGB object might be a string already or have a value
                            if isinstance(rgb_obj, str):
                                return rgb_obj
                            # Try to access as indexed (tuple-like)
                            if hasattr(rgb_obj, '__iter__') and not isinstance(rgb_obj, str):
                                try:
                                    return f"{rgb_obj[0]:02X}{rgb_obj[1]:02X}{rgb_obj[2]:02X}"
                                except:
                                    pass
                            # Convert to string as fallback
                            return str(rgb_obj) if rgb_obj else None
                        except:
                            return None
                    
                    # Safely get color values
                    font_color = None
                    if cell.font and cell.font.color:
                        try:
                            font_color = rgb_to_string(cell.font.color.rgb)
                        except:
                            pass
                    
                    bg_color = None
                    if cell.fill and hasattr(cell.fill, 'fgColor') and cell.fill.fgColor:
                        try:
                            bg_color = rgb_to_string(cell.fill.fgColor.rgb)
                        except:
                            pass
                    
                    cell_data = {
                        "value": cell.value,
                        "formatted_value": str(cell.value) if cell.value is not None else "",
                        "formula": cell.value if isinstance(cell.value, str) and cell.value.startswith('=') else None,
                        "style": {
                            "bold": cell.font.bold if cell.font else False,
                            "italic": cell.font.italic if cell.font else False,
                            "color": font_color,
                            "bg_color": bg_color
                        }
                    }
                    row_data.append(cell_data)
                data.append(row_data)
            
            return {
                "success": True,
                "file_path": file_path,
                "sheet_names": sheet_names,
                "active_sheet": ws.title,
                "data": data,
                "rows": ws.max_row,
                "columns": ws.max_column,
                "message": "Spreadsheet opened successfully"
            }
        except Exception as e:
            logger.error(f"Error opening spreadsheet: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def save_spreadsheet(self, file_path: str, data: Dict[str, List[List[Any]]] = None, new_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Save data to Excel spreadsheet with multiple sheets
        
        Args:
            file_path: Path to save the spreadsheet
            data: Dictionary with sheet names as keys and 2D list of cell values as values
            new_path: Optional new path to save as
        
        Returns:
            Dictionary with operation result
        """
        if not HAS_OPENPYXL:
            return {
                "success": False,
                "error": "openpyxl library not installed"
            }
        
        try:
            # Use new_path if provided, otherwise use file_path
            save_path = new_path if new_path else file_path
            
            # Ensure directory exists
            directory = os.path.dirname(save_path)
            if directory and not os.path.exists(directory):
                os.makedirs(directory, exist_ok=True)
            
            # Ensure file has .xlsx extension
            if not save_path.lower().endswith('.xlsx'):
                save_path = save_path + '.xlsx'
            
            # Load existing workbook or create new one
            if os.path.exists(file_path):
                wb = load_workbook(file_path)
                logger.info(f"Loaded existing workbook with sheets: {wb.sheetnames}")
            else:
                wb = Workbook()
                # Remove default sheet if we have custom data
                if data and wb.active and 'Sheet' in wb.active.title:
                    wb.remove(wb.active)
                logger.info("Created new workbook")
            
            # If data is provided, save all sheets
            if data:
                logger.info(f"Saving {len(data)} sheets: {list(data.keys())}")
                
                # Process each sheet
                for sheet_name, sheet_data in data.items():
                    # Create sheet if it doesn't exist
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                        # Clear existing data
                        for row in ws.iter_rows():
                            for cell in row:
                                cell.value = None
                        logger.info(f"Cleared existing sheet: {sheet_name}")
                    else:
                        ws = wb.create_sheet(sheet_name)
                        logger.info(f"Created new sheet: {sheet_name}")
                    
                    # Write data to sheet
                    rows_written = 0
                    cells_written = 0
                    for row_idx, row_data in enumerate(sheet_data, start=1):
                        if not row_data:  # Skip empty rows
                            continue
                        row_has_data = False
                        for col_idx, cell_value in enumerate(row_data, start=1):
                            if cell_value:  # Only write non-empty cells
                                cell = ws.cell(row=row_idx, column=col_idx)
                                
                                # Handle different value types
                                if isinstance(cell_value, dict):
                                    if cell_value.get('formula'):
                                        cell.value = cell_value['formula']
                                    else:
                                        cell.value = cell_value.get('value')
                                    
                                    # Apply formatting if provided
                                    style = cell_value.get('style', {})
                                    if style.get('bold') or style.get('italic'):
                                        cell.font = Font(bold=style.get('bold', False), 
                                                       italic=style.get('italic', False))
                                    
                                    if style.get('bg_color'):
                                        cell.fill = PatternFill(start_color=style['bg_color'], 
                                                               end_color=style['bg_color'], 
                                                               fill_type="solid")
                                else:
                                    cell.value = str(cell_value) if cell_value else None
                                
                                if cell.value:
                                    cells_written += 1
                                    row_has_data = True
                        
                        if row_has_data:
                            rows_written += 1
                    
                    logger.info(f"Sheet '{sheet_name}': wrote {rows_written} rows, {cells_written} cells")
            
            # Save the workbook
            wb.save(save_path)
            logger.info(f"Saved Excel spreadsheet to: {save_path}")
            
            return {
                "success": True,
                "file_path": save_path,
                "sheets": wb.sheetnames,
                "message": f"Spreadsheet saved successfully with {len(wb.sheetnames)} sheet(s)"
            }
        except Exception as e:
            logger.error(f"Error saving spreadsheet: {e}", exc_info=True)
            return {
                "success": False,
                "error": str(e)
            }
    
    async def add_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        Add a new sheet to an existing spreadsheet
        
        Args:
            file_path: Path to the spreadsheet
            sheet_name: Name of the new sheet
        
        Returns:
            Dictionary with operation result
        """
        if not HAS_OPENPYXL:
            return {
                "success": False,
                "error": "openpyxl library not installed"
            }
        
        try:
            if not os.path.exists(file_path):
                return {
                    "success": False,
                    "error": f"Spreadsheet not found: {file_path}"
                }
            
            wb = load_workbook(file_path)
            
            # Check if sheet already exists
            if sheet_name in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' already exists"
                }
            
            # Create new sheet
            wb.create_sheet(sheet_name)
            wb.save(file_path)
            
            logger.info(f"Added sheet '{sheet_name}' to {file_path}")
            return {
                "success": True,
                "message": f"Sheet '{sheet_name}' added successfully",
                "sheet_names": wb.sheetnames
            }
        except Exception as e:
            logger.error(f"Error adding sheet: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def delete_sheet(self, file_path: str, sheet_name: str) -> Dict[str, Any]:
        """
        Delete a sheet from a spreadsheet
        
        Args:
            file_path: Path to the spreadsheet
            sheet_name: Name of the sheet to delete
        
        Returns:
            Dictionary with operation result
        """
        if not HAS_OPENPYXL:
            return {
                "success": False,
                "error": "openpyxl library not installed"
            }
        
        try:
            if not os.path.exists(file_path):
                return {
                    "success": False,
                    "error": f"Spreadsheet not found: {file_path}"
                }
            
            wb = load_workbook(file_path)
            
            # Check if sheet exists
            if sheet_name not in wb.sheetnames:
                return {
                    "success": False,
                    "error": f"Sheet '{sheet_name}' not found"
                }
            
            # Can't delete the only sheet
            if len(wb.sheetnames) == 1:
                return {
                    "success": False,
                    "error": "Cannot delete the only sheet in the workbook"
                }
            
            # Delete sheet
            del wb[sheet_name]
            wb.save(file_path)
            
            logger.info(f"Deleted sheet '{sheet_name}' from {file_path}")
            return {
                "success": True,
                "message": f"Sheet '{sheet_name}' deleted successfully",
                "sheet_names": wb.sheetnames
            }
        except Exception as e:
            logger.error(f"Error deleting sheet: {e}")
            return {
                "success": False,
                "error": str(e)
            }
    
    async def cleanup(self):
        """Cleanup resources"""
        logger.info("Excel service cleanup completed")


# Global service instance
excel_service = ExcelService()

